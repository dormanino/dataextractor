from openpyxl import load_workbook
from Globus_Data.DataPoint import data_sbc_vehicles
from Globus_Data.Models.PurchasingData import PartCostData


class LoadGlobusData:

    @staticmethod
    def load_globus_data():
        wb = load_workbook(data_sbc_vehicles, read_only=True)
        ws = wb[wb.sheetnames[0]]
        globus_data_list = []
        index = 0
        for row in ws.rows:
            if index == 0:
                index += 1
                continue
            xablau = PartCostData(row[1].value, row[2].value, row[3].value, row[5].value, row[10].value, row[11].value,
                                  row[12].value, row[13].value, row[14].value, row[29].value, row[30].value,
                                  row[31].value)
            yield xablau
            globus_data_list.append(xablau)
        return globus_data_list
